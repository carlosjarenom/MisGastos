"""
services/excel.py — Import/Export de Excel
"""
import openpyxl
from models.schema import get_db


def import_excel(filepath: str) -> dict:
    """
    Importar Excel de Sonia a SQLite.
    Retorna: {"imported": N, "skipped": N, "errors": [...]}

    ⚠️ EN MANTENIMIENTO — El parser actual es frágil y produce datos incorrectos.
    Se está reescribiendo para soportar la estructura real del Excel de Sonia
    (gastos | ingresos, una hoja por mes, métodos de pago abreviados, etc.).
    """
    return {
        "imported": 0,
        "skipped": 0,
        "errors": ["Importación en mantenimiento. El parser se está reescribiendo para soportar el formato correcto del Excel."]
    }


def export_excel(month: str = None, year: str = None) -> str:
    """Exportar datos a Excel (formato compatible con Sonia).
    Retorna: path al archivo generado.
    """
    from config import BASE_DIR
    import os

    conn = get_db()
    c = conn.cursor()

    where = "WHERE kind='expense'"
    params = []
    if month and year:
        where += " AND date LIKE ?"
        params.append(f"{year}-{month}-%")

    c.execute(f"SELECT t.date, m.name, t.total, t.payment_method, c.name as category FROM transactions t LEFT JOIN merchants m ON t.merchant_id = m.id LEFT JOIN categories c ON t.category_id = c.id {where} ORDER BY t.date", params)
    rows = c.fetchall()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Gastos"

    # Headers
    ws.append(["Fecha", "Descripción", "Importe", "Método", "Categoría"])

    for row in rows:
        ws.append([
            row[0],
            row[1] or "",
            f"{row[2]:.2f}",
            row[3] or "",
            row[4] or "",
        ])

    filename = f"misgastos_{year}_{month}.xlsx"
    path = os.path.join(BASE_DIR, "data", filename)
    wb.save(path)

    conn.close()
    return path


def export_month_excel(year: str, month: str) -> str:
    """Exportar un mes concreto a Excel elegante y serio.

    Genera un archivo .xlsx con:
    - Cabecera con nombre del mes y año
    - Columnas con bordes y colores
    - Formato de moneda en euros
    - Total con fórmula SUM
    - Anchos de columna automáticos
    """
    import os
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from config import BASE_DIR

    meses_es = {'01': 'Enero', '02': 'Febrero', '03': 'Marzo', '04': 'Abril',
                '05': 'Mayo', '06': 'Junio', '07': 'Julio', '08': 'Agosto',
                '09': 'Septiembre', '10': 'Octubre', '11': 'Noviembre', '12': 'Diciembre'}

    month_name = meses_es.get(month, month)

    conn = get_db()
    c = conn.cursor()

    month_start = f"{year}-{month}-01"
    if month == "12":
        next_month_start = f"{int(year)+1}-01-01"
    else:
        next_month_start = f"{year}-{int(month)+1:02d}-01"

    c.execute("""
        SELECT t.date, t.description, m.name as merchant, t.total, t.payment_method,
            cat.name as category, t.card_last4, t.vehicle
        FROM transactions t
        LEFT JOIN merchants m ON t.merchant_id = m.id
        LEFT JOIN categories cat ON t.category_id = cat.id
        WHERE t.kind='expense' AND t.date >= ? AND t.date < ?
        ORDER BY t.date
    """, (month_start, next_month_start))
    rows = c.fetchall()
    conn.close()

    wb = Workbook()
    ws = wb.active
    ws.title = f"{month_name} {year}"

    # === ESTILOS ===
    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='1F2937', end_color='1F2937', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

    title_font = Font(name='Calibri', size=16, bold=True, color='1F2937')
    subtitle_font = Font(name='Calibri', size=11, color='6B7280')

    data_font = Font(name='Calibri', size=10)
    data_align = Alignment(vertical='center')

    total_font = Font(name='Calibri', size=11, bold=True, color='1F2937')
    total_fill = PatternFill(start_color='E5E7EB', end_color='E5E7EB', fill_type='solid')

    thin_border = Border(
        left=Side(style='thin', color='D1D5DB'),
        right=Side(style='thin', color='D1D5DB'),
        top=Side(style='thin', color='D1D5DB'),
        bottom=Side(style='thin', color='D1D5DB')
    )

    euro_format = '#,##0.00\\ "€"'

    # === TÍTULO ===
    ws.merge_cells('A1:H1')
    ws['A1'] = f"Gastos \u2014 {month_name} {year}"
    ws['A1'].font = title_font
    ws['A1'].alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[1].height = 28

    ws.merge_cells('A2:H2')
    ws['A2'] = f"Total de tickets: {len(rows)}"
    ws['A2'].font = subtitle_font
    ws.row_dimensions[2].height = 18

    # === CABECERAS (fila 4) ===
    headers = ['Fecha', 'Comercio', 'Descripción', 'Categoría', 'Total', 'Método de pago', 'Tarjeta', 'Coche']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
    ws.row_dimensions[4].height = 24

    # === DATOS (fila 5+) ===
    for row_idx, row in enumerate(rows, 5):
        values = [
            row['date'],
            row['merchant'] or '',
            row['description'] or '',
            row['category'] or '',
            float(row['total']) if row['total'] else 0,
            row['payment_method'] or '',
            f"****{row['card_last4']}" if row['card_last4'] else '',
            row['vehicle'] or ''
        ]
        for col, value in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.font = data_font
            cell.alignment = data_align
            cell.border = thin_border
            if col == 5:
                cell.number_format = euro_format

    # === TOTAL CON FÓRMULA ===
    total_row = len(rows) + 5
    ws.merge_cells(f'A{total_row}:D{total_row}')
    ws.cell(row=total_row, column=1, value='TOTAL').font = total_font
    ws.cell(row=total_row, column=1).alignment = Alignment(horizontal='right', vertical='center')
    ws.cell(row=total_row, column=1).fill = total_fill

    total_cell = ws.cell(row=total_row, column=5, value=f'=SUM(E5:E{total_row-1})')
    total_cell.font = total_font
    total_cell.fill = total_fill
    total_cell.number_format = euro_format
    total_cell.border = thin_border

    for col in range(1, 9):
        cell = ws.cell(row=total_row, column=col)
        cell.fill = total_fill
        cell.border = thin_border

    # === ANCHOS DE COLUMNA ===
    col_widths = [12, 25, 30, 15, 12, 15, 12, 15]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    # === FREEZAR PANELES ===
    ws.freeze_panes = 'A5'

    # === GUARDAR ===
    filename = f"gastos_{year}_{month}.xlsx"
    path = os.path.join(BASE_DIR, "data", filename)
    wb.save(path)

    return path
